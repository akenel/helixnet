"""BL-131 — the Migration Workbench: a catalog worklist as a real .xlsx an operator LIKES using.

Why a spreadsheet at all? Because the enrichment engine only existed as scripts a developer runs from
a terminal — the shop could never drive it. A workbook hands them the keys: export the unfinished rows,
walk the shelf filling in what only a human standing there knows (which variant, the shelf price, and
the BARCODE — scanned straight into the cell, since a scanner gun is just a keyboard), import it back,
and let the AI fill what a machine can find (image, description, specs).

Design rules (learned the hard way):
- **Formulas + data validation, never macros.** A .xlsm macro is blocked by default, dies in Google
  Sheets, and dies in LibreOffice. Formulas/dropdowns/conditional-formatting survive all three.
- **The category column is a DROPDOWN of canonical labels only.** The taxonomy funnel enforced at the
  point of entry — a free-text category can't be invented in the first place (BL-CAT doctrine).
- **A per-row Google link.** The operator's real workflow: search the name, land on the right page in
  ten seconds. Automate the click, not the judgement.
- **Live progress formulas.** The counter that makes a grind feel finite (same reason the test sheets
  carry a X/N counter and a stopwatch).
"""
from __future__ import annotations

import io
from datetime import date
from typing import Iterable, List, Optional
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.services.brand_registry import detect_brand, image_query, official_site, search_query
from src.services.catalog_taxonomy import CANONICAL_CATEGORIES

# La Piazza / Banco house colours (match the SOP + test-sheet artifacts)
RED = "C0392B"
DARK = "8B0000"
INK = "1F2937"
MUT = "6B7280"
LINE = "E5E7EB"
OK_BG = "E7F6EF"
WARN_BG = "FEF3E2"
BAD_BG = "FDEAEA"
HEAD_BG = "1F2937"

ACTIONS = ["ENRICH", "DONE", "SKIP", "DELETE"]

# (header, width, is_operator_editable)
COLUMNS = [
    ("Status",        11, False),   # formula
    ("SKU",           20, False),   # identity — never edit
    ("Product name",  46, True),
    ("Brand",          14, False),  # detected from the name — tells the operator which registry we matched
    ("Barcode / EAN", 18, True),    # ← scan the gun straight into this cell
    ("Category",      24, True),    # dropdown, canonical only
    ("Price",         11, True),   # header re-labelled with the store's currency at build time
    ("Cost",          11, True),
    ("Size / variant", 16, True),
    ("Photo?",         8, False),
    ("Text?",          8, False),
    ("We already have", 40, False),  # our own supplier catalog — ASK THIS BEFORE GOOGLE
    ("Look it up",    12, False),   # hyperlink → web search
    ("Find a photo",  12, False),   # hyperlink → image search (the operator picks the real one)
    ("Source URL",    34, True),
    ("Action",        12, True),    # dropdown
    ("Notes",         34, True),
]


def _style_header(ws, row=1):
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side("thin", color=RED))
    ws.row_dimensions[row].height = 26


def _title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=15, color=DARK)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(size=10, color=MUT)


def build_worklist_workbook(rows: Iterable[dict], *, section: Optional[str] = None,
                            env: str = "sandbox", stamp: Optional[str] = None,
                            currency: str = "CHF") -> bytes:
    """rows: dicts with sku/name/barcode/category/price/cost/size/has_image/has_text.

    Returns .xlsx bytes: START HERE · Worklist · Summary · Lists.
    """
    rows = list(rows)
    stamp = stamp or date.today().isoformat()
    wb = Workbook()

    # ---------------- Lists (feeds the dropdowns; kept last & hidden) ----------------
    lists = wb.create_sheet("Lists")
    lists["A1"] = "Categories (canonical — do not edit)"
    lists["A1"].font = Font(bold=True, size=10, color=MUT)
    for i, c in enumerate(CANONICAL_CATEGORIES, start=2):
        lists.cell(row=i, column=1, value=c)
    lists["C1"] = "Actions"
    lists["C1"].font = Font(bold=True, size=10, color=MUT)
    for i, a in enumerate(ACTIONS, start=2):
        lists.cell(row=i, column=3, value=a)
    lists.column_dimensions["A"].width = 30

    # ---------------- START HERE ----------------
    intro = wb.active
    intro.title = "START HERE"
    serial = (section or "ALL").upper().replace(" ", "-").replace("&", "AND")[:16]
    _title(intro, "📦 Banco — Catalog Worklist",
           f"🐺 La Piazza · № LP-WB-{stamp.replace('-','')}-{serial} · env: {env} · {len(rows)} products")
    guide = [
        ("", ""),
        ("How this works", ""),
        ("1.", "Open the Worklist tab. Each row is a product that still needs finishing."),
        ("2.", "Walk the shelf. Fill in ONLY what you can see: the exact variant, the shelf price, and the BARCODE."),
        ("3.", "💡 Your scanner gun is just a keyboard — click the Barcode cell and SCAN. It types straight in."),
        ("4.", "Don't know something? Leave it blank. Blank is fine — never guess."),
        ("5.", "Click '🔎 Look it up' on any row to Google it — grab the exact name in seconds."),
        ("6.", "Save the file and import it back into Banco. The AI fills in the picture and the description."),
        ("", ""),
        ("The rules", ""),
        ("⏱️", "If one product fights you for more than a minute — set Action = SKIP and move on. Never rabbit-hole."),
        ("🎯", "Category is a dropdown. Only real categories. If none fit, leave it and add a Note."),
        ("💰", "Price = what it sells for on the shelf. Cost can wait — leave it blank."),
        ("📷", "Photo? / Text? show what Banco already has. 'no' = the AI will go find it."),
        ("", ""),
        ("Which pile is this?", "(the 6 types — see the Catalog Playbook)"),
        ("FAST", "Branded consumables (papers, filters, lighters, e-liquids) — scan, confirm, done."),
        ("CONFIRM", "Variant families — same brand, 50 look-alikes. Read the pack, get the exact one."),
        ("SLOW", "Generic glass / CBD — no barcode, no catalog. Delivery slip → our own photo + label."),
        ("LINE", "Artisan (jewelry, oils) — don't name every piece. One line + options + price tiers."),
    ]
    r = 4
    for a, b in guide:
        intro.cell(row=r, column=1, value=a).font = Font(bold=True, size=10, color=DARK)
        intro.cell(row=r, column=2, value=b).font = Font(size=10.5, color=INK)
        intro.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    intro.column_dimensions["A"].width = 12
    intro.column_dimensions["B"].width = 104
    intro.sheet_view.showGridLines = False

    # ---------------- Worklist ----------------
    ws = wb.create_sheet("Worklist")
    for i, (h, w, _) in enumerate(COLUMNS, start=1):
        # Money headers carry THIS shop's currency — sandbox trades in EUR (Artemis Roma), so a
        # hardcoded "Price CHF" would ask the operator for the wrong money.
        ws.cell(row=1, column=i, value=(f"{h} {currency}" if h in ("Price", "Cost") else h))
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws)

    col = {h: get_column_letter(i) for i, (h, _, _) in enumerate(COLUMNS, start=1)}
    locked_fill = PatternFill("solid", fgColor="F3F4F6")

    for n, row in enumerate(rows, start=2):
        name = (row.get("name") or "").strip()
        ws[f"{col['SKU']}{n}"] = row.get("sku") or ""
        ws[f"{col['Product name']}{n}"] = name
        ws[f"{col['Barcode / EAN']}{n}"] = row.get("barcode") or ""
        ws[f"{col['Barcode / EAN']}{n}"].number_format = "@"      # text — never mangle a long EAN
        ws[f"{col['Category']}{n}"] = row.get("category") or ""
        if row.get("price") is not None:
            ws[f"{col['Price']}{n}"] = float(row["price"])
        ws[f"{col['Price']}{n}"].number_format = '#,##0.00'
        if row.get("cost") is not None:
            ws[f"{col['Cost']}{n}"] = float(row["cost"])
        ws[f"{col['Cost']}{n}"].number_format = '#,##0.00'
        ws[f"{col['Size / variant']}{n}"] = row.get("size") or ""
        ws[f"{col['Photo?']}{n}"] = "yes" if row.get("has_image") else "no"
        ws[f"{col['Text?']}{n}"] = "yes" if row.get("has_text") else "no"
        # What our OWN supplier catalog already knows. Shown BEFORE the Google links on purpose:
        # 10,284 FourTwenty rows (99% images, 100% prices) sat unused while the sheet sent the
        # operator to the web to find things we already own.
        ref = row.get("ref")
        if ref:
            c = ws[f"{col['We already have']}{n}"]
            price = f" · CHF {ref['price']}" if ref.get("price") else ""
            exact = "✅" if ref.get("score", 0) >= 0.85 else "≈"
            c.value = f"{exact} {ref['title'][:44]}{price}"
            c.fill = PatternFill("solid", fgColor=OK_BG if ref.get("score", 0) >= 0.85 else WARN_BG)
            c.font = Font(size=9.5, color=INK)

        # The operator's real workflow, one click each. Two links because they answer two different
        # questions: "what IS this exactly?" (web) and "which of these hundreds is the real pack?"
        # (images). The machine can search hundreds; only the person holding the box can pick the
        # right one — so we hand them the shortlist and let them choose.
        if name:
            link = ws[f"{col['Look it up']}{n}"]
            link.value = "🔎 Google"
            link.hyperlink = f"https://www.google.com/search?q={quote_plus(search_query(name))}"
            link.font = Font(color="0563C1", underline="single", size=10)

            # Scoped to the brand's official site when we know it → every image on that page IS the
            # real pack. The label says WHICH site, so the operator knows what they're looking at.
            brand = detect_brand(name)
            site = official_site(brand)
            img = ws[f"{col['Find a photo']}{n}"]
            img.value = f"🖼️ {site}" if site else "🖼️ Images"
            img.hyperlink = f"https://www.google.com/search?tbm=isch&q={quote_plus(image_query(name))}"
            img.font = Font(color="0563C1", underline="single", size=10)
            if brand:
                ws[f"{col['Brand']}{n}"] = brand
                ws[f"{col['Brand']}{n}"].fill = locked_fill

        ws[f"{col['Action']}{n}"] = "ENRICH"
        # Status formula: ready only when the human bits are in
        ws[f"{col['Status']}{n}"] = (
            f'=IF({col["Action"]}{n}="SKIP","⏭ skip",'
            f'IF(AND({col["Product name"]}{n}<>"",{col["Price"]}{n}>0,{col["Category"]}{n}<>""),"✅ ready",'
            f'IF({col["Product name"]}{n}="","❌ no name","⏳ needs work")))'
        )
        for h in ("Status", "SKU", "Photo?", "Text?"):
            ws[f"{col[h]}{n}"].fill = locked_fill
        for h in ("Photo?", "Text?", "Status"):
            ws[f"{col[h]}{n}"].alignment = Alignment(horizontal="center")

    last = max(len(rows) + 1, 2)

    # dropdowns — the funnel, enforced at the point of entry
    dv_cat = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${len(CANONICAL_CATEGORIES)+1}",
                            allow_blank=True, showDropDown=False)
    dv_cat.error = "Pick a category from the list — new categories are not allowed here."
    dv_cat.errorTitle = "Use the dropdown"
    ws.add_data_validation(dv_cat)
    dv_cat.add(f"{col['Category']}2:{col['Category']}{last}")

    dv_act = DataValidation(type="list", formula1=f"=Lists!$C$2:$C${len(ACTIONS)+1}",
                            allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_act)
    dv_act.add(f"{col['Action']}2:{col['Action']}{last}")

    # traffic lights on Status
    rng = f"{col['Status']}2:{col['Status']}{last}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("ready",{col["Status"]}2))'],
        fill=PatternFill("solid", fgColor=OK_BG), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("needs",{col["Status"]}2))'],
        fill=PatternFill("solid", fgColor=WARN_BG), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("no name",{col["Status"]}2))'],
        fill=PatternFill("solid", fgColor=BAD_BG), stopIfTrue=False))
    # highlight a missing price — the #1 thing only a human at the shelf can supply
    ws.conditional_formatting.add(f"{col['Price']}2:{col['Price']}{last}", FormulaRule(
        formula=[f'AND({col["Action"]}2<>"SKIP",OR({col["Price"]}2="",{col["Price"]}2=0))'],
        fill=PatternFill("solid", fgColor=WARN_BG), stopIfTrue=False))

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"

    # ---------------- Summary ----------------
    s = wb.create_sheet("Summary")
    _title(s, "📊 Progress", f"Live — it updates as you fill the Worklist · {stamp}")
    W = "Worklist"
    # Derived from the COLUMNS map, never hardcoded: adding one column would otherwise silently point
    # every counter at the wrong field — a Summary that lies is worse than no Summary.
    def rng(header):
        c = col[header]
        return f"{W}!{c}2:{c}{last}"
    stats = [
        ("Products in this list", f'=COUNTA({rng("SKU")})'),
        ("✅ Ready to import", f'=COUNTIF({rng("Status")},"*ready*")'),
        ("⏳ Still need work", f'=COUNTIF({rng("Status")},"*needs*")'),
        ("⏭ Skipped", f'=COUNTIF({rng("Status")},"*skip*")'),
        ("", ""),
        ("Barcodes captured 🔫", f'=COUNTA({rng("Barcode / EAN")})'),
        ("Prices filled 💰", f'=COUNTIF({rng("Price")},">0")'),
        ("Categories set 🗂", f'=COUNTA({rng("Category")})'),
        ("Pages linked 🔗", f'=COUNTA({rng("Source URL")})'),
        ("", ""),
        ("Photos Banco already has 📷", f'=COUNTIF({rng("Photo?")},"yes")'),
        ("Descriptions it already has ✍", f'=COUNTIF({rng("Text?")},"yes")'),
    ]
    r = 4
    for label, formula in stats:
        if label:
            s.cell(row=r, column=1, value=label).font = Font(size=11, color=INK)
            c = s.cell(row=r, column=2, value=formula)
            c.font = Font(size=13, bold=True, color=DARK)
            c.alignment = Alignment(horizontal="right")
        r += 1
    s.cell(row=r + 1, column=1, value="% ready").font = Font(size=11, bold=True, color=INK)
    pc = s.cell(row=r + 1, column=2, value=f'=IF(B4=0,0,B5/B4)')
    pc.number_format = "0%"
    pc.font = Font(size=18, bold=True, color=RED)
    pc.alignment = Alignment(horizontal="right")
    s.cell(row=r + 3, column=1,
           value="Work one shelf at a time. When this hits 100%, save and import it back into Banco.").font = \
        Font(size=10, italic=True, color=MUT)
    s.column_dimensions["A"].width = 34
    s.column_dimensions["B"].width = 14
    s.sheet_view.showGridLines = False

    wb.move_sheet("Lists", offset=3)
    lists.sheet_state = "hidden"
    wb.active = 0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------------------------
# READ-BACK (BL-131 step 2) — parse the workbook the operator filled at the shelf.
# ---------------------------------------------------------------------------------------------

class WorkbookError(Exception):
    """The uploaded file isn't a worklist we can read."""


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def parse_worklist_workbook(data: bytes) -> List[dict]:
    """Read a filled Worklist tab back into plain dicts (one per row).

    Deliberately forgiving about everything EXCEPT identity: we match the header row by NAME (so a
    reordered/hidden column can't silently shift the data into the wrong field), and a row without a
    SKU is skipped rather than guessed at. Values are returned RAW — validation, barcode-cleaning and
    category-funnelling happen at the import endpoint, where the DB is there to check against.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)   # data_only: read formula RESULTS
    except Exception as e:
        raise WorkbookError(f"Not a readable .xlsx file ({str(e)[:80]})")
    if "Worklist" not in wb.sheetnames:
        raise WorkbookError("No 'Worklist' tab — is this the workbook Banco exported?")
    ws = wb["Worklist"]

    # header name -> column index (never trust position; a user may hide/reorder columns)
    hdr = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if isinstance(h, str) and h.strip():
            hdr[h.strip().lower()] = c
    need = ["sku", "product name"]
    missing = [n for n in need if n not in hdr]
    if missing:
        raise WorkbookError(f"Missing column(s): {', '.join(missing)}")

    def col(name):
        return hdr.get(name)

    def col_starts(prefix):
        """Money headers carry the shop's currency ("Price EUR" / "Price CHF"), so they can't be
        matched by an exact name — a sheet exported from a EUR shop must still import."""
        for h, i in hdr.items():
            if h.startswith(prefix):
                return i
        return None

    out = []
    for r in range(2, ws.max_row + 1):
        sku = _cell(ws, r, col("sku"))
        if not sku:
            continue                      # blank row, or a row the operator added without identity
        row = {
            "row": r,
            "sku": str(sku),
            "name": _cell(ws, r, col("product name")),
            "barcode": _cell(ws, r, col("barcode / ean")) if col("barcode / ean") else None,
            "category": _cell(ws, r, col("category")) if col("category") else None,
            "price": _cell(ws, r, col_starts("price")) if col_starts("price") else None,
            "cost": _cell(ws, r, col_starts("cost")) if col_starts("cost") else None,
            "size": _cell(ws, r, col("size / variant")) if col("size / variant") else None,
            "source_url": _cell(ws, r, col("source url")) if col("source url") else None,
            "action": (_cell(ws, r, col("action")) or "ENRICH") if col("action") else "ENRICH",
            "notes": _cell(ws, r, col("notes")) if col("notes") else None,
        }
        # a gun/keyboard can leave a number typed as a float ("42425700.0") — normalise to a code
        if row["barcode"] is not None and not isinstance(row["barcode"], str):
            row["barcode"] = format(row["barcode"], ".0f") if float(row["barcode"]).is_integer() else str(row["barcode"])
        row["action"] = str(row["action"]).strip().upper()
        out.append(row)
    return out
